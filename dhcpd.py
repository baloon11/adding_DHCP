# coding: utf-8
import os
from openpyxl import load_workbook


if os.path.exists('dhcpd.conf'):
    os.remove('dhcpd.conf')

xlsx_line_list=[]
wb =load_workbook('Network Addresses.xlsx', use_iterators=True)
ws = wb.get_sheet_by_name('Sheet1')
for row in ws.iter_rows():
    curr_row=[]
    for cell in row:
        curr_row.append(cell.value)
    xlsx_line_list.append(curr_row)
xlsx_line_list=xlsx_line_list[1:]

mac_list=[]
ip_list=[]
hostname_list=[]

def separate(sep):
    mac_list.append(str(sep[0]))
    ip_list.append(str(sep[1]))
    hostname_list.append(str(sep[2]))
map(separate, xlsx_line_list)

def create_block(ip='',mac='',hostname=''):
    block='host '+hostname+ '{\n\
        \thardware ethernet '+mac+';\n\
        \tfixed-address '+ip+';\n\
        \toption host-name "CAN1";\n\
        \toption option-150 tftp_address;\n\
        \toption BITI.transfer-mode "tftp";\n\
        \toption BITI.config-file-name "'+hostname+'.confg";\n}\n\n'
    return block

for index in xrange(len(ip_list)):
    out=open('dhcpd.conf','a')
    out.write(create_block(ip=ip_list[index],\
                            mac=mac_list[index],\
                            hostname=hostname_list[index]))
    out.close()